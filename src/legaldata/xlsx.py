from datetime import datetime

import duckdb
from openpyxl import Workbook
from openpyxl.styles import Font

ORDEM_TABELAS = ["processos", "movimentos", "assuntos", "comunicacoes", "partes", "advogados"]

TABELAS_INFO = {
    "processos":    ("id", "Dados principais do processo (classe, órgão, grau, datas)"),
    "movimentos":   ("id, movimento_seq", "Histórico de movimentação do processo"),
    "assuntos":     ("id, assunto_seq", "Assuntos do processo"),
    "comunicacoes": ("id", "Publicações do DJEN (texto limpo, sem HTML)"),
    "partes":       ("comunicacao_id, parte_seq", "Partes das comunicações DJEN (polo A/P)"),
    "advogados":    ("comunicacao_id, advogado_seq", "Advogados das comunicações DJEN (número e UF da OAB)"),
}


def _tabelas_existentes(con: duckdb.DuckDBPyConnection) -> list[str]:
    existentes = {r[0] for r in con.execute(
        "SELECT table_name FROM information_schema.tables WHERE table_schema='main'"
    ).fetchall()}
    return [t for t in ORDEM_TABELAS if t in existentes]


def _write_table_sheet(wb: Workbook, con: duckdb.DuckDBPyConnection, tabela: str) -> int:
    cursor = con.execute(f"SELECT * FROM {tabela}")
    colunas = [d[0] for d in cursor.description]
    linhas = cursor.fetchall()

    ws = wb.create_sheet(tabela)
    ws.append(colunas)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for linha in linhas:
        ws.append(list(linha))
    ws.freeze_panes = "A2"

    return len(linhas)


def _write_orientacoes_sheet(wb: Workbook, db_path: str, contagens: dict[str, int]) -> None:
    ws = wb.create_sheet("Orientações", 0)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 10

    ws.append(["legaldata — exportação XLSX"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Banco de origem: {db_path}"])
    ws.append([f"Gerado em: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.append([])

    header_row = ws.max_row + 1
    ws.append(["Tabela", "Chave primária", "Descrição", "Linhas"])
    for cell in ws[header_row]:
        cell.font = Font(bold=True)

    for tabela in ORDEM_TABELAS:
        if tabela in contagens:
            pk, desc = TABELAS_INFO[tabela]
            ws.append([tabela, pk, desc, contagens[tabela]])


def export_xlsx(db_path: str, output_path: str) -> dict[str, int]:
    """Exporta todas as tabelas de um banco DuckDB para um único XLSX,
    uma aba por tabela mais uma aba "Orientações"."""
    con = duckdb.connect(db_path, read_only=True)
    tabelas = _tabelas_existentes(con)

    wb = Workbook()
    wb.remove(wb.active)

    contagens: dict[str, int] = {}
    for tabela in tabelas:
        contagens[tabela] = _write_table_sheet(wb, con, tabela)

    con.close()

    _write_orientacoes_sheet(wb, db_path, contagens)
    wb.save(output_path)

    return contagens
